"""
build_compliance_crosswalk.py

Builds Week3_Compliance_Crosswalk.xlsx: maps every Week 3 control to
ISO/IEC 27001:2022 Annex A, the Microsoft Cloud Security Benchmark
(MCSB v1 - the current production standard; MCSB v2 is in preview,
not GA as of mid-2026, so v1 is cited throughout for reproducibility),
and the DPDP Act 2023 + DPDP Rules 2025 (notified 14 Nov 2025).

Run once to produce the workbook; kept in the repo so the crosswalk
can be regenerated/extended (e.g. when MCSB v2 goes GA) without
hand-editing cell-by-cell.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_PATH = "Week3_Compliance_Crosswalk.xlsx"

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT_NAME, size=14, bold=True, color="1F3864")
SUBTITLE_FONT = Font(name=FONT_NAME, size=10, italic=True, color="595959")
BODY_FONT = Font(name=FONT_NAME, size=10)
BOLD_BODY = Font(name=FONT_NAME, size=10, bold=True)
WRAP = Alignment(wrap_text=True, vertical="top", horizontal="left")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BAND_FILL = PatternFill("solid", fgColor="F2F5FA")

wb = Workbook()

# ============================================================
# Sheet 1: Compliance Crosswalk
# ============================================================
ws = wb.active
ws.title = "Compliance Crosswalk"

ws["A1"] = "CloudGuardian - Week 3 Compliance Crosswalk"
ws["A1"].font = TITLE_FONT
ws["A2"] = ("Maps every Week 3 remediation / governance control to ISO/IEC 27001:2022 Annex A, "
            "the Microsoft Cloud Security Benchmark (MCSB v1), and the DPDP Act 2023 + DPDP Rules 2025.")
ws["A2"].font = SUBTITLE_FONT
ws.merge_cells("A1:J1")
ws.merge_cells("A2:J2")
ws.row_dimensions[2].height = 30
ws["A2"].alignment = WRAP

HEADERS = [
    "Control ID",
    "Week 3 Component",
    "Control Description",
    "ISO/IEC 27001:2022\nAnnex A Ref(s)",
    "ISO Annex A Control Name",
    "MCSB v1\nControl ID",
    "MCSB Control Name",
    "DPDP Act 2023 / Rules 2025 Provision",
    "Automation Level",
    "Evidence / How Demonstrated",
]

header_row = 4
for col, text in enumerate(HEADERS, start=1):
    c = ws.cell(row=header_row, column=col, value=text)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = BORDER

ROWS = [
    ["storage_public_access", "Function: remediations/storage_public_access.py",
     "Disable account-level anonymous blob public access (allow_blob_public_access=False).",
     "A.8.9; A.8.12; A.5.10",
     "Configuration management; Data leakage prevention; Acceptable use of assets",
     "DP-2; NS-2", "Discover/classify/protect sensitive data; Secure cloud services with network controls",
     "Sec 8(5) reasonable security safeguards - access control",
     "Auto (Low/Med) or Approval-gated (High/Crit)",
     "Function execution log in Log Analytics; before/after state in audit record"],

    ["storage_encryption", "Function: remediations/storage_encryption.py",
     "Enforce HTTPS-only transport and minimum TLS 1.2 on the target Storage Account.",
     "A.8.24; A.8.20",
     "Use of cryptography; Networks security",
     "DP-3; DP-4", "Encrypt sensitive data in transit; Enable data-at-rest encryption by default",
     "DPDP Rules 2025 Sch. - encryption as a listed reasonable security safeguard",
     "Auto (Low/Med) or Approval-gated (High/Crit)",
     "Function execution log; Prowler/ScoutSuite re-scan clears the finding"],

    ["diagnostic_logging", "Function: remediations/diagnostic_logging.py",
     "Re-attach a Diagnostic Setting sending logs+metrics to the Log Analytics Workspace.",
     "A.8.15; A.8.16",
     "Logging; Monitoring activities",
     "LT-3; LT-4", "Enable logging for security investigation; Enable network logging",
     "Rules 2025: retain logs >=1 year to support breach detection/investigation",
     "Auto (Low/Med) or Approval-gated (High/Crit)",
     "Diagnostic setting visible in Portal; KQL query against Log Analytics"],

    ["sql_encryption", "Function: remediations/sql_encryption.py",
     "Enable Transparent Data Encryption on the Azure SQL Database.",
     "A.8.24",
     "Use of cryptography",
     "DP-4", "Enable data-at-rest encryption by default",
     "DPDP Rules 2025 Sch. - encryption as a listed reasonable security safeguard",
     "Auto (Low/Med) or Approval-gated (High/Crit)",
     "transparentDataEncryptions.get() state=Enabled; audit record"],

    ["keyvault_firewall", "Function: remediations/keyvault_firewall.py",
     "Set Key Vault network ACL default_action=Deny with AzureServices bypass + explicit allow-list.",
     "A.8.20; A.8.9",
     "Networks security; Configuration management",
     "NS-2; DP-8", "Secure cloud services with network controls; Ensure security of key/cert repository",
     "Sec 8(5) reasonable security safeguards - access control",
     "Approval-gated only (network change risk - see code comment)",
     "vaults.get() network_acls.default_action=Deny; audit record"],

    ["tagging", "Function: remediations/tagging.py",
     "Merge required governance tags (Environment, Owner, DataClassification, CostCenter).",
     "A.5.9; A.5.10",
     "Inventory of information and other associated assets; Acceptable use of assets",
     "GS-1", "Define asset management and data protection strategy",
     "Supports records-of-processing / data classification obligations",
     "Auto (Low/Med) or Approval-gated (High/Crit)",
     "Resource tags visible in Portal/Resource Graph; audit record"],

    ["approval_gate", "Logic App: logic_app/approval_workflow.json",
     "Human-in-the-loop Approve/Reject gate for High/Critical findings before any remediation runs.",
     "A.5.15; A.8.32",
     "Access control; Change management",
     "PA-7", "Follow just enough administration principle",
     "Demonstrates accountability principle underpinning Sec 8 obligations",
     "N/A - this IS the approval control",
     "Logic App run history (who approved/rejected, when)"],

    ["least_privilege_identity", "Terraform: rbac.tf (custom role + Managed Identity)",
     "Function App and Automation Account authenticate via System-Assigned Managed Identity under a "
     "custom role scoped to only the actions the 6 remediations need - not built-in Contributor.",
     "A.5.15; A.8.2; A.8.5",
     "Access control; Privileged access rights; Secure authentication",
     "PA-1; IM-3", "Separate and limit highly privileged/administrative users; Manage application identities securely",
     "Sec 8(5) reasonable security safeguards - access control & accountability",
     "N/A - this is the control enabling all others",
     "azurerm_role_definition + role_assignment in Terraform state; Portal IAM blade"],

    ["governance_drift_check", "Automation Account runbook: Test-RemediationDrift.ps1",
     "Daily scheduled re-check of all 6 controls; alerts (does not silently re-fix) on drift.",
     "A.5.35; A.5.36; A.8.16",
     "Independent review of information security; Compliance with policies/standards; Monitoring activities",
     "LT-5; GS-3", "Centralize security log management and analysis; Define and implement security posture management",
     "Supports ongoing 'reasonable security safeguards' obligation, not just point-in-time",
     "Auto (detection only, never auto-fixes)",
     "Automation job history + JobLogs/JobStreams in Log Analytics"],

    ["privacy_preserving_llm", "privacy_llm/tokenizer.py + llm_client.py",
     "Pseudonymizes subscription IDs, object IDs, UPNs, resource names, principal names before any "
     "LLM call; hard-blocks the call if a leakage guardrail check fails; detokenizes only after the "
     "response returns inside the Azure boundary.",
     "A.8.11; A.8.12; A.5.34",
     "Data masking; Data leakage prevention; Privacy and protection of PII",
     "DP-2", "Discover, classify, and protect sensitive data",
     "Rules 2025 Sch.: explicitly lists 'encryption, obfuscation, masking, or use of virtual tokens "
     "mapped to personal data' as a reasonable security safeguard - this control implements that "
     "safeguard literally, for UPNs (personal data under the Act's definition)",
     "N/A - this is a preventive control, not a remediation",
     "verify_no_leakage() raise on failure; unit test in privacy_llm/ (see setup guide Step 14)"],
]

r = header_row + 1
for i, row in enumerate(ROWS):
    for col, value in enumerate(row, start=1):
        c = ws.cell(row=r, column=col, value=value)
        c.font = BODY_FONT
        c.alignment = WRAP
        c.border = BORDER
        if i % 2 == 1:
            c.fill = BAND_FILL
    ws.cell(row=r, column=1).font = BOLD_BODY
    ws.row_dimensions[r].height = 60
    r += 1

# Summary block using real formulas (not hardcoded), per house style.
summary_row = r + 1
ws.cell(row=summary_row, column=1, value="Total controls mapped:").font = BOLD_BODY
ws.cell(row=summary_row, column=2, value=f"=COUNTA(A{header_row+1}:A{r-1})").font = BODY_FONT

ws.cell(row=summary_row + 1, column=1, value="Auto-remediable (no approval needed):").font = BOLD_BODY
ws.cell(row=summary_row + 1, column=2,
        value=f'=COUNTIF(I{header_row+1}:I{r-1},"Auto (Low/Med) or Approval-gated (High/Crit)")').font = BODY_FONT

ws.cell(row=summary_row + 2, column=1, value="Approval-gated only (network-risk controls):").font = BOLD_BODY
ws.cell(row=summary_row + 2, column=2,
        value=f'=COUNTIF(I{header_row+1}:I{r-1},"Approval-gated only (network change risk - see code comment)")').font = BODY_FONT

col_widths = [16, 30, 42, 16, 30, 12, 30, 34, 22, 34]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A5"

# ============================================================
# Sheet 2: Legend & Methodology
# ============================================================
ws2 = wb.create_sheet("Legend & Methodology")
ws2["A1"] = "Legend & Methodology"
ws2["A1"].font = TITLE_FONT
ws2.column_dimensions["A"].width = 26
ws2.column_dimensions["B"].width = 100

legend_rows = [
    ("ISO/IEC 27001:2022 Annex A",
     "93 controls across 4 themes: A.5 Organizational (37), A.6 People (8), A.7 Physical (14), "
     "A.8 Technological (34). This is the 2022 revision (superseded the 2013/114-control version) - "
     "cite '27001:2022' explicitly in your report and defense, reviewers will check this."),
    ("Microsoft Cloud Security Benchmark (MCSB)",
     "MCSB v1 is the current GA/production standard as of this document (mid-2026) and is what's "
     "cited throughout. MCSB v2 is in public preview (adds an AI Security domain) but has not reached "
     "General Availability - do not cite v2 control IDs as authoritative yet. MCSB is the rebrand of "
     "the earlier 'Azure Security Benchmark' (renamed October 2022); Defender for Cloud's Regulatory "
     "Compliance dashboard maps directly onto MCSB control IDs, so these IDs are also what you'll see "
     "if you enable that dashboard for your own screenshot evidence."),
    ("DPDP Act 2023 / DPDP Rules 2025",
     "The Digital Personal Data Protection Act received assent in August 2023; the implementing DPDP "
     "Rules, 2025 were notified on 14 November 2025, giving the Act full operational effect. "
     "Implementation is PHASED: the Data Protection Board was instituted from 13 Nov 2025; Consent "
     "Manager registration begins 12 months later (13 Nov 2026); the main compliance duties - notice "
     "requirements, security safeguards, breach notification, Significant Data Fiduciary obligations, "
     "Data Principal rights - become enforceable at 18 months (13 May 2027). Frame this in your report "
     "as 'designed for compliance ahead of the mandatory date,' not 'currently legally required' - that "
     "is the factually accurate and more impressive framing."),
    ("Automation Level column",
     "'Auto' = routed by Event Grid directly to the Function App, no human involved (Low/Medium "
     "severity only, by design - see event_grid.tf). 'Approval-gated' = routed to the Logic App, "
     "requires a human Approve click before the Function executes (High/Critical severity). "
     "keyvault_firewall is approval-gated ONLY (never auto) even at low severity, because a firewall "
     "tightening is the one control in this set that can plausibly break a legitimate caller - see the "
     "code comment in remediations/keyvault_firewall.py for the full rationale."),
    ("How to defend this table",
     "For each row, be ready to show: (1) the Terraform/Python file implementing it, (2) a live Portal "
     "screenshot of the after-state, (3) the audit log entry in Log Analytics. This table is the index "
     "linking all three - walk your evaluator across one row end-to-end as your 'proof of traceability' "
     "answer if asked how you know the mapping isn't just paperwork."),
]
row = 3
for title, body in legend_rows:
    ws2.cell(row=row, column=1, value=title).font = BOLD_BODY
    ws2.cell(row=row, column=1).alignment = WRAP
    c = ws2.cell(row=row, column=2, value=body)
    c.font = BODY_FONT
    c.alignment = WRAP
    ws2.row_dimensions[row].height = 90
    row += 2

wb.save(OUT_PATH)
print(f"Saved {OUT_PATH}")
